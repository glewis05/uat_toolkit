# reporters/excel_export.py
# ============================================================================
# UAT RESULTS EXCEL EXPORT
# ============================================================================
# PURPOSE: Export UAT cycle results to Excel for stakeholders.
#
# EXCEL SHEETS GENERATED:
#   - Summary: Cycle metadata and progress metrics
#   - All Tests: Full test list with results
#   - By Tester: Progress breakdown by tester
#   - Failed Tests: Tests that failed (for dev handoff)
#   - Retest Queue: Tests needing retest with dev status
#
# AVIATION ANALOGY:
#   Like generating a mission debrief report with all the details
#   needed for post-mission review and lessons learned.
#
# ============================================================================

import os
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from database.db_manager import UATDatabase


# Style definitions
HEADER_FILL = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
PASS_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FAIL_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
BLOCKED_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def export_uat_results(
    cycle_id: str,
    output_dir: str = None,
    output_filename: str = None,
    db_path: str = None
) -> Dict:
    """
    PURPOSE:
        Export cycle results to Excel for stakeholders.

    R EQUIVALENT:
        Like writexl::write_xlsx() with multiple sheets from a list of tibbles.

    PARAMETERS:
        cycle_id (str): Cycle to export
        output_dir (str): Output directory (default: outputs/)
        output_filename (str): Output filename (auto-generated if not provided)
        db_path (str): Optional database path

    RETURNS:
        dict: {
            'success': bool,
            'file_path': str,
            'sheets': ['Summary', 'All Tests', ...],
            'test_count': int
        }

    WHY THIS APPROACH:
        - Multiple sheets for different audiences
        - Conditional formatting shows pass/fail at a glance
        - Includes all data needed for stakeholder review
    """
    if not OPENPYXL_AVAILABLE:
        return {
            'success': False,
            'error': 'openpyxl not installed. Run: pip install openpyxl'
        }

    # Determine output path
    if not output_dir:
        output_dir = Path(__file__).parent.parent / 'outputs'
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if not output_filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"UAT_Results_{cycle_id}_{timestamp}.xlsx"

    output_path = output_dir / output_filename

    # Get data
    with UATDatabase(db_path) as db:
        cycle = db.get_cycle(cycle_id)
        if not cycle:
            return {'success': False, 'error': f'Cycle not found: {cycle_id}'}

        gate_status = db.get_gate_status(cycle_id)
        tester_progress = db.get_tester_progress(cycle_id)
        retest_queue = db.get_retest_queue(cycle_id)

        # Get all test cases
        conn = db.get_connection()
        cursor = conn.execute("""
            SELECT * FROM uat_test_cases
            WHERE uat_cycle_id = ?
            ORDER BY profile_id, test_id
        """, (cycle_id,))
        all_tests = [dict(row) for row in cursor.fetchall()]

    # Create workbook
    wb = Workbook()

    # ========== SUMMARY SHEET ==========
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_data = [
        ['UAT CYCLE RESULTS'],
        [''],
        ['Cycle ID', cycle['cycle_id']],
        ['Name', cycle['name']],
        ['Type', cycle['uat_type']],
        ['Status', cycle['status']],
        ['Program', cycle.get('program_name', 'N/A')],
        ['Clinical PM', cycle.get('clinical_pm', 'N/A')],
        ['Target Launch', cycle.get('target_launch_date', 'N/A')],
        [''],
        ['TEST PROGRESS'],
        ['Total Tests', cycle.get('total_tests') or 0],
        ['Passed', cycle.get('passed') or 0],
        ['Failed', cycle.get('failed') or 0],
        ['Blocked', cycle.get('blocked') or 0],
        ['Not Run', cycle.get('not_run') or 0],
        ['Execution %', f"{cycle.get('execution_pct') or 0}%"],
        ['Pass Rate', f"{cycle.get('pass_rate') or 0}%"],
        [''],
        ['PRE-UAT GATE'],
        ['Items Complete', f"{gate_status.get('completed') or 0}/{gate_status.get('total') or 0}"],
        ['Gate Passed', 'Yes' if cycle.get('pre_uat_gate_passed') else 'No'],
        [''],
        ['GO/NO-GO DECISION'],
        ['Decision', cycle.get('go_nogo_decision', 'Pending')],
        ['Signed By', cycle.get('go_nogo_signed_by', 'N/A')],
        ['Date', cycle.get('go_nogo_signed_date', 'N/A')],
    ]

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx in (1, 11, 20, 24):  # Section headers
                cell.font = Font(bold=True, size=12)

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 40

    # ========== ALL TESTS SHEET ==========
    ws_tests = wb.create_sheet("All Tests")

    test_headers = [
        'Test ID', 'Profile ID', 'Title', 'Platform', 'Change ID',
        'Target Rule', 'Test Type', 'Status', 'Tested By', 'Tested Date',
        'Notes', 'Defect ID'
    ]

    for col_idx, header in enumerate(test_headers, 1):
        cell = ws_tests.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for row_idx, test in enumerate(all_tests, 2):
        values = [
            test.get('test_id'),
            test.get('profile_id'),
            test.get('title'),
            test.get('platform'),
            test.get('change_id'),
            test.get('target_rule'),
            test.get('test_type'),
            test.get('test_status'),
            test.get('tested_by'),
            test.get('tested_date'),
            test.get('execution_notes'),
            test.get('defect_id')
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws_tests.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER

            # Status-based highlighting
            if col_idx == 8:  # Status column
                if value == 'Pass':
                    cell.fill = PASS_FILL
                elif value == 'Fail':
                    cell.fill = FAIL_FILL
                elif value == 'Blocked':
                    cell.fill = BLOCKED_FILL

    # Auto-width columns
    for col_idx, header in enumerate(test_headers, 1):
        ws_tests.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 15)

    # ========== BY TESTER SHEET ==========
    ws_testers = wb.create_sheet("By Tester")

    tester_headers = ['Tester', 'Assignment Type', 'Total', 'Completed', 'Passed', 'Failed', 'Blocked', 'Not Run', 'Completion %']

    for col_idx, header in enumerate(tester_headers, 1):
        cell = ws_testers.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for row_idx, tp in enumerate(tester_progress, 2):
        values = [
            tp.get('assigned_to'),
            tp.get('assignment_type'),
            tp.get('total_tests'),
            tp.get('completed'),
            tp.get('passed'),
            tp.get('failed'),
            tp.get('blocked'),
            tp.get('not_run'),
            f"{tp.get('completion_pct') or 0}%"
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws_testers.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER

    for col_idx, header in enumerate(tester_headers, 1):
        ws_testers.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 12)

    # ========== FAILED TESTS SHEET ==========
    ws_failed = wb.create_sheet("Failed Tests")

    failed_tests = [t for t in all_tests if t.get('test_status') == 'Fail']

    failed_headers = [
        'Test ID', 'Profile ID', 'Title', 'Platform', 'Target Rule',
        'Defect ID', 'Defect Description', 'Dev Status', 'Dev Notes',
        'Tested By', 'Notes'
    ]

    for col_idx, header in enumerate(failed_headers, 1):
        cell = ws_failed.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for row_idx, test in enumerate(failed_tests, 2):
        values = [
            test.get('test_id'),
            test.get('profile_id'),
            test.get('title'),
            test.get('platform'),
            test.get('target_rule'),
            test.get('defect_id'),
            test.get('defect_description'),
            test.get('dev_status'),
            test.get('dev_notes'),
            test.get('tested_by'),
            test.get('execution_notes')
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws_failed.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.fill = FAIL_FILL

    for col_idx, header in enumerate(failed_headers, 1):
        ws_failed.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 15)

    # ========== RETEST QUEUE SHEET ==========
    ws_retest = wb.create_sheet("Retest Queue")

    retest_headers = [
        'Test ID', 'Profile ID', 'Title', 'Platform', 'Target Rule',
        'Initial Status', 'Initial Tester', 'Defect ID', 'Dev Status', 'Dev Notes',
        'Retest Status', 'Retest By', 'Retest Date', 'Retest Notes'
    ]

    for col_idx, header in enumerate(retest_headers, 1):
        cell = ws_retest.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for row_idx, test in enumerate(retest_queue, 2):
        values = [
            test.get('test_id'),
            test.get('profile_id'),
            test.get('title'),
            test.get('platform'),
            test.get('target_rule'),
            test.get('initial_status'),
            test.get('initial_tester'),
            test.get('defect_id'),
            test.get('dev_status'),
            test.get('dev_notes'),
            test.get('retest_status'),
            test.get('retest_by'),
            test.get('retest_date'),
            test.get('retest_notes')
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws_retest.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER

    for col_idx, header in enumerate(retest_headers, 1):
        ws_retest.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 15)

    # Save workbook
    wb.save(str(output_path))

    return {
        'success': True,
        'file_path': str(output_path),
        'sheets': ['Summary', 'All Tests', 'By Tester', 'Failed Tests', 'Retest Queue'],
        'test_count': len(all_tests),
        'failed_count': len(failed_tests),
        'retest_count': len(retest_queue)
    }
