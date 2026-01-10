# importers/nccn_importer.py
# ============================================================================
# NCCN UAT PACKAGE IMPORTER
# ============================================================================
# PURPOSE: Import NCCN test profiles from Excel UAT packages.
#
# NCCN UAT packages contain:
#   - Test Profile Catalog: All test profiles with conditions
#   - Tester 1, Tester 2, etc.: Per-tester assignment sheets
#
# TEST TYPE MAPPING:
#   POS → 'positive' (scenario SHOULD trigger rule)
#   NEG → 'negative' (scenario should NOT trigger rule)
#   DEP → 'deprecated' (rule was removed)
#
# AVIATION ANALOGY:
#   Like importing a mission package with test scenarios.
#   Each profile is a specific flight condition to validate.
#
# ============================================================================

import os
import uuid
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from database.db_manager import UATDatabase


# Column mapping for NCCN test profile catalog
# Maps Excel column headers to database fields
NCCN_COLUMN_MAP = {
    'Profile ID': 'profile_id',
    'Test Profile ID': 'profile_id',
    'Change ID': 'change_id',
    'Rule ID': 'target_rule',
    'NCCN Rule': 'target_rule',
    'Change Type': 'change_type',
    'Platform': 'platform',
    'Test Type': 'test_type',
    'Type': 'test_type',
    'Patient Conditions': 'patient_conditions',
    'Conditions': 'patient_conditions',
    'Expected Outcome': 'expected_results',
    'Expected Result': 'expected_results',
    'Cross Trigger': 'cross_trigger_check',
    'Notes': 'notes',
}

# Test type normalization
TEST_TYPE_MAP = {
    'POS': 'positive',
    'POSITIVE': 'positive',
    'NEG': 'negative',
    'NEGATIVE': 'negative',
    'DEP': 'deprecated',
    'DEPRECATED': 'deprecated',
}


def import_nccn_profiles(
    file_path: str,
    cycle_id: str,
    sheet_name: str = "Test Profile Catalog",
    program_id: str = None,
    preview_only: bool = True,
    db_path: str = None
) -> Dict:
    """
    PURPOSE:
        Import NCCN test profiles from Excel into uat_test_cases.

    R EQUIVALENT:
        Like readxl::read_excel() piped to purrr::pmap() for insertion.

    PARAMETERS:
        file_path (str): Path to NCCN UAT Package Excel file
        cycle_id (str): UAT cycle to associate profiles with
        sheet_name (str): Sheet to import (default: "Test Profile Catalog")
        program_id (str): Program ID (resolved from cycle if not provided)
        preview_only (bool): If True, show what would happen without inserting

    EXPECTED EXCEL COLUMNS:
        Profile ID, Change ID, Rule ID, Change Type, Platform,
        Test Type (POS/NEG/DEP), Patient Conditions, Expected Outcome

    RETURNS:
        dict: {
            'profiles_found': int,
            'profiles_created': int,
            'profiles_updated': int,
            'by_platform': {'P4M': n, 'Px4M': n},
            'by_change_type': {'NEW': n, 'MODIFIED': n},
            'by_test_type': {'positive': n, 'negative': n},
            'errors': [...]
        }

    WHY THIS APPROACH:
        - Column mapping handles variations in header names
        - Test type normalization (POS→positive) for consistency
        - Preview mode prevents accidental imports
        - Full audit trail for Part 11 compliance
    """
    if not OPENPYXL_AVAILABLE:
        return {
            'error': 'openpyxl not installed. Run: pip install openpyxl',
            'profiles_found': 0,
            'profiles_created': 0
        }

    # Expand path and validate
    file_path = os.path.expanduser(file_path)
    if not os.path.exists(file_path):
        return {'error': f'File not found: {file_path}', 'profiles_found': 0}

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        return {'error': f'Failed to open Excel file: {e}', 'profiles_found': 0}

    if sheet_name not in wb.sheetnames:
        return {
            'error': f'Sheet "{sheet_name}" not found. Available: {wb.sheetnames}',
            'profiles_found': 0
        }

    ws = wb[sheet_name]

    # Get headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value else '')

    # Map headers to field names
    field_indices = {}
    for idx, header in enumerate(headers):
        if header in NCCN_COLUMN_MAP:
            field_name = NCCN_COLUMN_MAP[header]
            field_indices[field_name] = idx

    # Verify required fields
    required = ['profile_id', 'test_type']
    missing = [f for f in required if f not in field_indices]
    if missing:
        return {
            'error': f'Missing required columns: {missing}. Found: {list(field_indices.keys())}',
            'profiles_found': 0
        }

    # Parse rows
    profiles = []
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Extract values by field
        profile_data = {}
        for field_name, col_idx in field_indices.items():
            cell_value = row[col_idx].value
            profile_data[field_name] = str(cell_value).strip() if cell_value else None

        # Skip empty rows
        if not profile_data.get('profile_id'):
            continue

        # Normalize test type
        raw_type = profile_data.get('test_type', '').upper()
        profile_data['test_type'] = TEST_TYPE_MAP.get(raw_type, raw_type.lower())

        # Generate test_id if not present
        if not profile_data.get('test_id'):
            profile_data['test_id'] = profile_data['profile_id']

        # Generate title from components
        if not profile_data.get('title'):
            parts = [
                profile_data.get('target_rule', ''),
                profile_data.get('test_type', ''),
                profile_data.get('platform', '')
            ]
            profile_data['title'] = ' - '.join(p for p in parts if p)

        profiles.append(profile_data)

    # Summarize by categories
    by_platform = {}
    by_change_type = {}
    by_test_type = {}

    for p in profiles:
        platform = p.get('platform', 'Unknown')
        by_platform[platform] = by_platform.get(platform, 0) + 1

        change_type = p.get('change_type', 'Unknown')
        by_change_type[change_type] = by_change_type.get(change_type, 0) + 1

        test_type = p.get('test_type', 'Unknown')
        by_test_type[test_type] = by_test_type.get(test_type, 0) + 1

    result = {
        'profiles_found': len(profiles),
        'by_platform': by_platform,
        'by_change_type': by_change_type,
        'by_test_type': by_test_type,
        'errors': errors,
        'preview_only': preview_only
    }

    if preview_only:
        result['profiles_created'] = 0
        result['profiles_updated'] = 0
        result['message'] = f"Preview: Would import {len(profiles)} profiles. Run with preview_only=False to import."
        return result

    # Import to database
    with UATDatabase(db_path) as db:
        conn = db.get_connection()

        # Get program_id from cycle if not provided
        if not program_id:
            cursor = conn.execute(
                "SELECT program_id FROM uat_cycles WHERE cycle_id = ?",
                (cycle_id,)
            )
            row = cursor.fetchone()
            if row:
                program_id = row['program_id']

        created = 0
        updated = 0

        for profile in profiles:
            test_id = profile['test_id']

            # Check if exists
            cursor = conn.execute(
                "SELECT test_id FROM uat_test_cases WHERE test_id = ?",
                (test_id,)
            )
            exists = cursor.fetchone() is not None

            if exists:
                # Update existing
                conn.execute("""
                    UPDATE uat_test_cases SET
                        uat_cycle_id = ?,
                        profile_id = ?,
                        platform = ?,
                        change_id = ?,
                        target_rule = ?,
                        change_type = ?,
                        test_type = ?,
                        patient_conditions = ?,
                        expected_results = ?,
                        cross_trigger_check = ?,
                        notes = COALESCE(?, notes),
                        updated_date = CURRENT_TIMESTAMP
                    WHERE test_id = ?
                """, (
                    cycle_id,
                    profile.get('profile_id'),
                    profile.get('platform'),
                    profile.get('change_id'),
                    profile.get('target_rule'),
                    profile.get('change_type'),
                    profile.get('test_type'),
                    profile.get('patient_conditions'),
                    profile.get('expected_results'),
                    profile.get('cross_trigger_check'),
                    profile.get('notes'),
                    test_id
                ))
                updated += 1
            else:
                # Insert new
                conn.execute("""
                    INSERT INTO uat_test_cases (
                        test_id, program_id, uat_cycle_id, title,
                        profile_id, platform, change_id, target_rule, change_type,
                        test_type, patient_conditions, expected_results,
                        cross_trigger_check, notes, test_status
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Not Run')
                """, (
                    test_id,
                    program_id,
                    cycle_id,
                    profile.get('title'),
                    profile.get('profile_id'),
                    profile.get('platform'),
                    profile.get('change_id'),
                    profile.get('target_rule'),
                    profile.get('change_type'),
                    profile.get('test_type'),
                    profile.get('patient_conditions'),
                    profile.get('expected_results'),
                    profile.get('cross_trigger_check'),
                    profile.get('notes')
                ))
                created += 1

        conn.commit()

        # Audit log
        db.log_audit(
            'uat_cycle', cycle_id, 'Profiles Imported',
            new_val=f"{created} created, {updated} updated from {os.path.basename(file_path)}",
            changed_by='nccn_importer',
            reason=f"NCCN profile import from {sheet_name}"
        )

    result['profiles_created'] = created
    result['profiles_updated'] = updated
    result['message'] = f"Imported {created} new, updated {updated} existing profiles."

    return result


def import_nccn_assignments(
    file_path: str,
    cycle_id: str,
    tester_sheet: str,
    tester_email: str,
    assignment_type: str = 'primary',
    preview_only: bool = True,
    db_path: str = None
) -> Dict:
    """
    PURPOSE:
        Import tester assignments from tester-specific sheets.

    PARAMETERS:
        file_path (str): Path to NCCN UAT Package Excel file
        cycle_id (str): UAT cycle ID
        tester_sheet (str): Sheet name, e.g., "Tester 1", "Tester 2"
        tester_email (str): Tester's email for assigned_to field
        assignment_type (str): 'primary', 'secondary', 'cross_check'
        preview_only (bool): If True, show what would happen

    RETURNS:
        dict: {
            'profiles_found': int,
            'assignments_made': int,
            'by_platform': {'P4M': n, ...}
        }

    WHY THIS APPROACH:
        - Separates profile import from assignment
        - Supports overlapping assignments for cross-validation
        - Preserves existing test data, only updates assignment fields
    """
    if not OPENPYXL_AVAILABLE:
        return {'error': 'openpyxl not installed', 'profiles_found': 0}

    file_path = os.path.expanduser(file_path)
    if not os.path.exists(file_path):
        return {'error': f'File not found: {file_path}', 'profiles_found': 0}

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        return {'error': f'Failed to open Excel file: {e}', 'profiles_found': 0}

    if tester_sheet not in wb.sheetnames:
        return {
            'error': f'Sheet "{tester_sheet}" not found. Available: {wb.sheetnames}',
            'profiles_found': 0
        }

    ws = wb[tester_sheet]

    # Find profile ID column
    headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
    profile_id_idx = None

    for idx, header in enumerate(headers):
        if header in ('Profile ID', 'Test Profile ID', 'profile_id'):
            profile_id_idx = idx
            break

    if profile_id_idx is None:
        return {'error': 'Profile ID column not found in sheet', 'profiles_found': 0}

    # Collect profile IDs
    profile_ids = []
    for row in ws.iter_rows(min_row=2):
        cell_value = row[profile_id_idx].value
        if cell_value:
            profile_ids.append(str(cell_value).strip())

    result = {
        'profiles_found': len(profile_ids),
        'tester': tester_email,
        'assignment_type': assignment_type,
        'preview_only': preview_only
    }

    if preview_only:
        result['assignments_made'] = 0
        result['message'] = f"Preview: Would assign {len(profile_ids)} profiles to {tester_email}"
        return result

    # Make assignments
    with UATDatabase(db_path) as db:
        conn = db.get_connection()
        assigned = 0

        for profile_id in profile_ids:
            # Find test by profile_id in cycle
            cursor = conn.execute("""
                UPDATE uat_test_cases SET
                    assigned_to = ?,
                    assignment_type = ?,
                    updated_date = CURRENT_TIMESTAMP
                WHERE uat_cycle_id = ? AND (profile_id = ? OR test_id = ?)
            """, (tester_email, assignment_type, cycle_id, profile_id, profile_id))

            if cursor.rowcount > 0:
                assigned += cursor.rowcount

        conn.commit()

        # Audit log
        db.log_audit(
            'uat_cycle', cycle_id, 'Tests Assigned',
            new_val=f"{assigned} tests assigned to {tester_email} ({assignment_type})",
            changed_by='nccn_importer',
            reason=f"Bulk assignment from {tester_sheet}"
        )

    result['assignments_made'] = assigned
    result['message'] = f"Assigned {assigned} tests to {tester_email}"

    return result
